#!/usr/bin/env python3
"""
Grupo MF CRM — Generador de datos seed
========================================
Convierte el xlsx de Grupo MF en data.js para que el CRM
los cargue automáticamente al abrir index.html.

Uso:
    python3 seed-data.py

Archivo requerido en la misma carpeta:
    - "../grupo-mf-workspace/DATABASE-GrupoMF _ Leads.xlsx"

Archivo generado:
    - data.js  (cargado automáticamente por index.html)
"""

import json
import hashlib
import os
import sys
from datetime import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "..", "grupo-mf-workspace", "DATABASE-GrupoMF _ Leads.xlsx")
OUTPUT_JS = os.path.join(HERE, "data.js")

try:
    import openpyxl
except ImportError:
    print("  ✗ Necesitás openpyxl: pip3 install openpyxl")
    sys.exit(1)


# ── Helpers ──────────────────────────────────────────────────────────

def uid_from(s: str) -> str:
    return hashlib.md5(s.encode()).hexdigest()[:16]

def fmt_phone(p):
    if p is None:
        return ""
    if isinstance(p, float):
        return str(int(p))
    return str(p).strip()

def fmt_date(d):
    if d is None:
        return None
    if isinstance(d, datetime):
        return d.strftime("%Y-%m-%d")
    try:
        return datetime.strptime(str(d).strip(), "%d/%m/%Y").strftime("%Y-%m-%d")
    except ValueError:
        return None

def normalize_tipo(t):
    if not t:
        return ""
    t = str(t).strip().upper()
    TIPO_MAP = {
        "COMERCIO / PYME": "PYME",
        "PYME": "PYME",
        "CADENA DE TIENDAS": "CADENA DE TIENDAS",
        "SUPERMERCADO": "SUPERMERCADO",
        "INTEGRADOR / PARTNER": "INTEGRADOR / PARTNER",
        "FARMACIA": "FARMACIA",
        "PERFUMERÍA": "PERFUMERÍA",
    }
    return TIPO_MAP.get(t, t)

def normalize_calif(c):
    if not c:
        return ""
    c = str(c).strip()
    if "muy calificado" in c.lower() or c.upper().startswith("A"):
        return "A"
    if "potencial" in c.lower() or c.upper().startswith("B"):
        return "B"
    if "frío" in c.lower() or c.upper().startswith("C"):
        return "C"
    return c


# ── Parsers ──────────────────────────────────────────────────────────

def parse_leads(path: str) -> list:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Leads | 2026"]
    leads = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1] or not str(row[1]).strip():
            continue
        fecha = fmt_date(row[0])
        nombre = str(row[1]).strip().upper()
        tipo = normalize_tipo(row[2])
        provincia = str(row[3] or "").strip().upper()
        ciudad = str(row[4] or "").strip().upper()
        mail = str(row[5] or "").strip()
        wpp = fmt_phone(row[6] if len(row) > 6 else None)
        calif = normalize_calif(row[7] if len(row) > 7 else "")
        notas = str(row[8] or "").strip() if len(row) > 8 else ""

        uid = uid_from(f"{nombre}|{fecha}|{wpp}")

        leads.append({
            "id": uid,
            "fecha_ingreso": fecha or datetime.now().strftime("%Y-%m-%d"),
            "nombre": nombre,
            "tipo_cliente": tipo,
            "calificacion": calif,
            "canal": "",  # Mora fills this
            "provincia": provincia,
            "ciudad": ciudad,
            "mail": mail,
            "wpp": wpp,
            "estado": "NUEVO",
            "fecha_ultimo_contacto": None,
            "dias_followup": 14,
            "notas": notas,
            "historial": [],
        })

    wb.close()
    return leads


def parse_ventas(path: str) -> list:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Ventas | 2026"]
    ventas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        fecha = fmt_date(row[0])
        if not fecha:
            continue
        nombre = str(row[1]).strip().upper()
        tipo = normalize_tipo(row[2])
        canal = str(row[3] or "").strip().upper()
        total_ars = float(row[4]) if row[4] and isinstance(row[4], (int, float)) else 0
        primera_raw = str(row[5] or "").strip().upper()
        primera = primera_raw in ("SÍ", "SI", "1", "TRUE")
        total_usd = float(row[6]) if len(row) > 6 and row[6] and isinstance(row[6], (int, float)) else 0

        uid = uid_from(f"{nombre}|{fecha}|{total_ars}")

        ventas.append({
            "id": uid,
            "fecha_venta": fecha,
            "nombre": nombre,
            "tipo_cliente": tipo,
            "canal": canal or "GOOGLE ADS",
            "total_ars": total_ars,
            "total_usd": total_usd,
            "primera_venta": primera,
            "notas": "",
        })

    wb.close()
    return ventas


# ── Main ─────────────────────────────────────────────────────────────

def main():
    print("Grupo MF CRM — Generador de seed data")
    print("=" * 42)

    if not os.path.exists(XLSX):
        print(f"  ✗ No encontrado: {XLSX}")
        sys.exit(1)

    print(f"  ✓ Source: {os.path.basename(XLSX)}")
    print()
    print("Procesando...")
    leads = parse_leads(XLSX)
    ventas = parse_ventas(XLSX)
    print(f"  Leads:  {len(leads):,}")
    print(f"  Ventas: {len(ventas):,}")

    # Build data.js
    leads_json = json.dumps(leads, ensure_ascii=False, separators=(",", ":"))
    ventas_json = json.dumps(ventas, ensure_ascii=False, separators=(",", ":"))
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")

    output = f"""// Grupo MF CRM — Seed Data
// Generado: {ts}
// Leads: {len(leads):,} | Ventas: {len(ventas):,}
//
// IMPORTANTE: Este archivo es generado automáticamente.
// No lo edites a mano. Para actualizar:
//   python3 seed-data.py

window.GMF_SEED = {{
  leads: {leads_json},
  ventas: {ventas_json}
}};
"""

    with open(OUTPUT_JS, "w", encoding="utf-8") as f:
        f.write(output)

    size_kb = len(output) / 1024
    print(f"\n  ✓ data.js generado ({size_kb:.0f} KB)")
    print(f"  ✓ Listo — abrí index.html para usar el CRM")
    print()


if __name__ == "__main__":
    main()
